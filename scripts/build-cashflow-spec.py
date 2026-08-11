"""Genera il file Excel della struttura di riclassificazione cash flow WEISS.

Legge il piano dei conti v4, ci innesta il mastro 40, assegna famiglia e
sottogruppo a ogni voce e scrive sei fogli: MAPPATURA, CASH FLOW, BUDGET,
SCOSTAMENTO, PER CENTRO, SINTESI.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import Outline

SRC = "/Users/nicolascarpa/Desktop/accounting/docs/Piano_dei_conti_gestionale_WEISS_v4.xlsx"
DST = "/Users/nicolascarpa/Desktop/accounting/docs/Cash_flow_riclassificazione_WEISS_v1.xlsx"

MESI = ["GEN", "FEB", "MAR", "APR", "MAG", "GIU",
        "LUG", "AGO", "SET", "OTT", "NOV", "DIC"]

# ---------------------------------------------------------------- mastro 40
MASTRO_40 = [
    ("40.1.01", "40 - Movimenti finanziari e patrimoniali", "40.1 - Investimenti",
     "Acquisto immobilizzazioni materiali", "Obbligatorio", ""),
    ("40.1.02", "40 - Movimenti finanziari e patrimoniali", "40.1 - Investimenti",
     "Acquisto immobilizzazioni immateriali", "Obbligatorio", ""),
    ("40.1.03", "40 - Movimenti finanziari e patrimoniali", "40.1 - Investimenti",
     "Migliorie su beni di terzi", "Obbligatorio", ""),
    ("40.1.04", "40 - Movimenti finanziari e patrimoniali", "40.1 - Investimenti",
     "Cessione cespiti", "Obbligatorio", "Incasso della vendita; la plusvalenza (12.07) non e' cassa"),
    ("40.2.01", "40 - Movimenti finanziari e patrimoniali", "40.2 - Finanziamenti",
     "Rimborso capitale mutui", "Default STR", "Solo quota capitale; gli interessi restano in 32.1.01"),
    ("40.2.02", "40 - Movimenti finanziari e patrimoniali", "40.2 - Finanziamenti",
     "Erogazione nuovi finanziamenti", "Default STR", ""),
    ("40.2.03", "40 - Movimenti finanziari e patrimoniali", "40.2 - Finanziamenti",
     "Rimborso finanziamento soci", "Default STR", ""),
    ("40.2.04", "40 - Movimenti finanziari e patrimoniali", "40.2 - Finanziamenti",
     "Versamento finanziamento soci", "Default STR", ""),
    ("40.3.01", "40 - Movimenti finanziari e patrimoniali", "40.3 - Erario e previdenza",
     "F24 IVA versata", "Default STR", ""),
    ("40.3.02", "40 - Movimenti finanziari e patrimoniali", "40.3 - Erario e previdenza",
     "F24 imposte sul reddito", "Default STR", "Versamento di IRES e IRAP; il mastro 33 resta competenza"),
    ("40.3.03", "40 - Movimenti finanziari e patrimoniali", "40.3 - Erario e previdenza",
     "F24 ritenute e contributi", "Default STR", "Ricompone il lordo del personale insieme a 28.1"),
    ("40.3.04", "40 - Movimenti finanziari e patrimoniali", "40.3 - Erario e previdenza",
     "Rimborsi e crediti compensati", "Default STR", ""),
    ("40.4.01", "40 - Movimenti finanziari e patrimoniali", "40.4 - Tesoreria interna",
     "Versamento contanti in banca", "Default STR", "Neutro: le due gambe si elidono nel consolidato"),
    ("40.4.02", "40 - Movimenti finanziari e patrimoniali", "40.4 - Tesoreria interna",
     "Giroconti tra conti", "Default STR", "Neutro: le due gambe si elidono nel consolidato"),
]

# --------------------------------------------------- struttura del prospetto
# (famiglia, titolo, [(sottogruppo, titolo, [codici])])
STRUTTURA = [
    ("A", "INCASSI OPERATIVI", [
        ("A1", "Corrispettivi", ["10.01", "10.09"]),
        ("A2", "Eventi", ["11.01", "11.02"]),
        ("A3", "Altri proventi", ["12.01", "12.02", "12.03", "12.04", "12.06", "13.01", "13.02"]),
    ]),
    ("B", "COSTO DEL VENDUTO", [
        ("B1", "Beverage alcolico", ["20.1.01", "20.1.02", "20.1.03", "20.1.04", "20.1.05"]),
        ("B2", "Beverage analcolico", ["20.2.01", "20.2.02", "20.2.03", "20.2.04"]),
        ("B3", "Caffetteria", ["20.3.01", "20.3.02", "20.3.03", "20.3.04"]),
        ("B4", "Food", ["20.4.01", "20.4.02", "20.4.03", "20.4.04", "20.4.05"]),
        ("B5", "Consumabili di servizio", ["20.5.01", "20.5.02", "20.5.03", "20.5.04", "20.5.05"]),
        ("B6", "Rettifiche su acquisti", ["20.6.01"]),
    ]),
    ("C", "COSTO DEL PERSONALE", [
        ("C1", "Retribuzioni", ["28.1.01", "28.1.02", "28.1.03", "28.1.04", "28.1.05",
                                "28.1.06", "28.1.07", "28.1.08", "28.1.09"]),
        ("C2", "Oneri sociali", ["28.2.01", "28.2.02", "28.2.03"]),
        ("C3", "TFR corrisposto", ["28.3.02"]),
        ("C4", "Altri costi del personale", ["28.4.01", "28.4.02", "28.4.03", "28.4.04", "28.4.05"]),
        ("C5", "Organi sociali e collaborazioni", ["29.01", "29.02", "29.04", "29.05"]),
    ]),
    ("D", "COSTI DIRETTI EVENTI", [
        ("D1", "Artisti e service", ["26.01", "26.02"]),
        ("D2", "Manodopera evento", ["26.03", "26.05", "26.06"]),
        ("D3", "Promozione evento", ["26.04", "26.07", "26.08"]),
        ("D4", "Oneri e allestimenti evento", ["26.09", "26.10", "26.11"]),
    ]),
    ("E", "COSTI DI STRUTTURA", [
        ("E1", "Immobili e spazi", ["27.01", "27.02", "27.03"]),
        ("E2", "Utenze", ["22.01", "22.02", "22.03", "22.04", "22.05", "22.07"]),
        ("E3", "Noleggi, leasing e licenze", ["27.04", "27.05", "27.06", "27.07", "27.08"]),
        ("E4", "Manutenzioni e servizi operativi", ["23.01", "23.02", "23.03", "23.05", "23.07"]),
        ("E5", "Attrezzatura e allestimenti", ["21.01", "21.02", "21.03", "21.04", "21.05", "21.06", "21.07"]),
        ("E6", "Servizi professionali e amministrativi",
         ["24.01", "24.02", "24.03", "24.04", "24.05", "24.06", "24.07", "24.08", "24.09"]),
        ("E7", "Marketing e comunicazione",
         ["25.01", "25.02", "25.03", "25.04", "25.05", "25.06", "25.07", "25.08"]),
        ("E8", "Tributi, assicurazioni e oneri diversi",
         ["30.01", "30.02", "30.03", "30.04", "30.05", "30.06", "30.07", "30.08", "30.09",
          "30.10", "30.13", "30.14", "30.15"]),
    ]),
    ("F", "ONERI FINANZIARI", [
        ("F1", "Interessi passivi", ["32.1.01", "32.1.02", "32.1.03", "32.1.04"]),
        ("F2", "Spese e servizi bancari", ["32.2.01", "32.2.02", "32.2.03", "32.2.04"]),
        ("F3", "Commissioni su incassi", ["32.3.01", "32.3.02", "32.3.03", "32.3.04", "32.3.05"]),
    ]),
    ("G", "FISCO E IVA", [
        ("G1", "IVA incassata sui corrispettivi", []),
        ("G2", "IVA pagata sugli acquisti", []),
        ("G3", "F24 IVA", ["40.3.01", "40.3.04"]),
        ("G4", "Imposte sul reddito", ["40.3.02"]),
        ("G5", "Ritenute e contributi", ["40.3.03"]),
    ]),
    ("H", "INVESTIMENTI", [
        ("H1", "Acquisto immobilizzazioni", ["40.1.01", "40.1.02", "40.1.03"]),
        ("H2", "Cessione cespiti", ["40.1.04"]),
    ]),
    ("I", "FINANZIAMENTI", [
        ("I1", "Rimborso capitale", ["40.2.01"]),
        ("I2", "Nuova finanza", ["40.2.02"]),
        ("I3", "Soci", ["40.2.03", "40.2.04"]),
    ]),
]

MEMO = ["40.4.01", "40.4.02"]

# voci che non toccano mai il conto corrente
FUORI_CASSA = {
    "12.07": "Plusvalenza contabile; l'incasso della cessione e' 40.1.04",
    "20.6.02": "Variazione di magazzino",
    "20.6.03": "Variazione di magazzino",
    "20.6.04": "Riclassifica di valore, nessun esborso",
    "20.6.05": "Riclassifica di valore, nessun esborso",
    "28.3.01": "Competenza; l'esborso e' 28.3.02",
    "30.11": "Mancata entrata, non un'uscita",
    "30.12": "Minusvalenza contabile",
    "31.01": "Non tocca il conto", "31.02": "Non tocca il conto",
    "31.03": "Non tocca il conto", "31.04": "Non tocca il conto",
    "31.05": "Non tocca il conto", "31.06": "Non tocca il conto",
    "31.07": "Non tocca il conto",
    "33.01": "Competenza; il versamento e' 40.3.02",
    "33.02": "Competenza; il versamento e' 40.3.02",
    "33.03": "Competenza; il versamento e' 40.3.02",
}

# entrate e rettifiche; tutto il resto e' uscita
ENTRATE = {"10.01", "11.01", "11.02", "12.01", "12.02", "12.03", "12.04", "12.06",
           "13.01", "13.02", "40.1.04", "40.2.02", "40.2.04", "40.3.04"}
RETTIFICHE = {"10.09": "Segno negativo: rettifica dei ricavi",
              "20.6.01": "Segno negativo: rettifica dei costi"}
NEUTRE = {"40.4.01", "40.4.02"}

# ------------------------------------------------------------------- stile
F_TITOLO = Font(name="Calibri", size=15, bold=True, color="111827")
F_SOTTOTIT = Font(name="Calibri", size=10, italic=True, color="6B7280")
F_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
F_FAM = Font(name="Calibri", size=11, bold=True, color="111827")
F_SOTTO = Font(name="Calibri", size=10, bold=True, color="374151")
F_VOCE = Font(name="Calibri", size=10, color="6B7280")
F_TOT = Font(name="Calibri", size=11, bold=True, color="7C2D12")
F_MEMO = Font(name="Calibri", size=10, italic=True, color="6B7280")
F_MEMO_B = Font(name="Calibri", size=10, italic=True, bold=True, color="374151")

FILL_HEADER = PatternFill("solid", fgColor="374151")
FILL_FAM = PatternFill("solid", fgColor="E5E7EB")
FILL_TOT = PatternFill("solid", fgColor="FEF3C7")
FILL_INPUT = PatternFill("solid", fgColor="FFFDF5")

BORDO_TOP = Border(top=Side(style="thin", color="9CA3AF"))
NUM = '#,##0.00;[Red]-#,##0.00;"-"'
PCT = '0.0%;[Red]-0.0%;"-"'


def leggi_piano():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Piano dei conti"]
    voci = {}
    ordine = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        codice, sezione, mastro, gruppo, nome, regola, note = (list(r) + [""] * 7)[:7]
        voci[codice] = {
            "codice": codice, "sezione": sezione, "mastro": mastro,
            "gruppo": gruppo or "", "nome": nome, "regola": regola,
            "note": note or "",
        }
        ordine.append(codice)
    for codice, mastro, gruppo, nome, regola, note in MASTRO_40:
        voci[codice] = {
            "codice": codice, "sezione": "PATRIMONIALE", "mastro": mastro,
            "gruppo": gruppo, "nome": nome, "regola": regola, "note": note,
        }
        ordine.append(codice)
    return voci, ordine


def arricchisci(voci):
    """Assegna famiglia, sottogruppo, segno e flag cassa."""
    for v in voci.values():
        v.update(famiglia="", fam_nome="", sottogruppo="", sot_nome="",
                 cassa="SI", segno="Uscita")
    for fam, fam_nome, sottogruppi in STRUTTURA:
        for sot, sot_nome, codici in sottogruppi:
            for c in codici:
                if c not in voci:
                    raise KeyError(f"codice {c} assente dal piano")
                voci[c].update(famiglia=fam, fam_nome=fam_nome,
                               sottogruppo=sot, sot_nome=sot_nome)
    for c in MEMO:
        voci[c].update(famiglia="MEMO", fam_nome="Tesoreria interna",
                       sottogruppo="M3", sot_nome="Tesoreria interna")
    for c, motivo in FUORI_CASSA.items():
        voci[c].update(cassa="NO", segno="-", famiglia="", fam_nome="",
                       sottogruppo="", sot_nome="")
        nota = voci[c]["note"]
        voci[c]["note"] = f"{nota} | Fuori cash flow: {motivo}".strip(" |")
    for c in voci:
        if voci[c]["cassa"] == "NO":
            continue
        if c in ENTRATE:
            voci[c]["segno"] = "Entrata"
        elif c in RETTIFICHE:
            voci[c]["segno"] = "Rettifica"
        elif c in NEUTRE:
            voci[c]["segno"] = "Neutro"
    return voci


# ------------------------------------------------------------- MAPPATURA
def foglio_mappatura(wb, voci, ordine):
    ws = wb.create_sheet("MAPPATURA")
    ws["A1"] = "MAPPATURA DELLE VOCI DI CONTO SUL CASH FLOW"
    ws["A1"].font = F_TITOLO
    ws["A2"] = ("Fonte di verita' della riclassificazione. Un movimento si registra sempre "
                "sulla voce di dettaglio: famiglia e sottogruppo sono derivati, mai scelti a mano.")
    ws["A2"].font = F_SOTTOTIT

    intestazioni = ["Codice", "Sezione", "Mastro", "Gruppo", "Voce di conto",
                    "Famiglia", "Sottogruppo", "Nome sottogruppo",
                    "Segno", "Cassa", "Regola centro di costo", "Note"]
    for i, h in enumerate(intestazioni, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font, c.fill = F_HEADER, FILL_HEADER
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[4].height = 22

    r = 5
    for codice in ordine:
        v = voci[codice]
        valori = [v["codice"], v["sezione"], v["mastro"], v["gruppo"], v["nome"],
                  v["famiglia"], v["sottogruppo"], v["sot_nome"],
                  v["segno"], v["cassa"], v["regola"], v["note"]]
        for i, val in enumerate(valori, 1):
            cell = ws.cell(row=r, column=i, value=val)
            cell.font = Font(name="Calibri", size=10,
                             color="9CA3AF" if v["cassa"] == "NO" else "111827")
            cell.alignment = Alignment(vertical="top", wrap_text=(i == 12))
        r += 1

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:L{r - 1}"
    for col, w in zip("ABCDEFGHIJKL",
                      [10, 15, 42, 32, 48, 10, 13, 34, 11, 8, 22, 60]):
        ws.column_dimensions[col].width = w
    return ws


# ------------------------------------------------------- prospetto (righe)
def scrivi_prospetto(ws, voci, titolo, sottotitolo, con_formule=True):
    """Scrive la scaletta gerarchica. Torna la mappa delle righe chiave."""
    ws.sheet_properties.outlinePr = Outline(summaryBelow=False, summaryRight=False)
    ws["A1"] = titolo
    ws["A1"].font = F_TITOLO
    ws["A2"] = sottotitolo
    ws["A2"].font = F_SOTTOTIT

    intest = ["Codice", "Voce"] + MESI + ["TOTALE"]
    for i, h in enumerate(intest, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font, c.fill = F_HEADER, FILL_HEADER
        c.alignment = Alignment(horizontal="center" if i > 2 else "left",
                                vertical="center")
    ws.row_dimensions[4].height = 22
    ws.freeze_panes = "C5"
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 46
    for i in range(3, 16):
        ws.column_dimensions[get_column_letter(i)].width = 13

    righe = {}
    r = 5

    def cella_num(row, col, formula=None, stile="voce"):
        c = ws.cell(row=row, column=col)
        if formula:
            c.value = formula
        c.number_format = NUM
        if stile == "voce":
            c.font, c.fill = F_VOCE, FILL_INPUT
        elif stile == "sotto":
            c.font = F_SOTTO
        elif stile == "fam":
            c.font, c.fill = F_FAM, FILL_FAM
        elif stile == "tot":
            c.font, c.fill, c.border = F_TOT, FILL_TOT, BORDO_TOP
        return c

    for fam, fam_nome, sottogruppi in STRUTTURA:
        r_fam = r
        ws.cell(row=r, column=1, value=fam).font = F_FAM
        ws.cell(row=r, column=1).fill = FILL_FAM
        ws.cell(row=r, column=2, value=fam_nome).font = F_FAM
        ws.cell(row=r, column=2).fill = FILL_FAM
        r += 1
        primo_dettaglio = r

        for sot, sot_nome, codici in sottogruppi:
            r_sot = r
            ws.cell(row=r, column=1, value=sot).font = F_SOTTO
            ws.cell(row=r, column=2, value="   " + sot_nome).font = F_SOTTO
            ws.row_dimensions[r].outlineLevel = 1
            r += 1
            r_voci_da = r
            for c in codici:
                ws.cell(row=r, column=1, value=c).font = F_VOCE
                ws.cell(row=r, column=2, value="      " + voci[c]["nome"]).font = F_VOCE
                ws.row_dimensions[r].outlineLevel = 2
                ws.row_dimensions[r].hidden = True
                for col in range(3, 16):
                    cella_num(r, col, stile="voce")
                ws.cell(row=r, column=15).value = f"=SUM(C{r}:N{r})"
                r += 1
            r_voci_a = r - 1

            for col in range(3, 16):
                L = get_column_letter(col)
                if codici:
                    f = f"=SUBTOTAL(9,{L}{r_voci_da}:{L}{r_voci_a})"
                    cella_num(r_sot, col, f, stile="sotto")
                else:
                    # G1 e G2: righe calcolate dai vatAmount, input diretto
                    c = cella_num(r_sot, col, stile="sotto")
                    c.fill = FILL_INPUT
            if not codici:
                ws.cell(row=r_sot, column=15).value = f"=SUM(C{r_sot}:N{r_sot})"
            righe[sot] = r_sot

        ultimo_dettaglio = r - 1
        for col in range(3, 16):
            L = get_column_letter(col)
            f = f"=SUBTOTAL(9,{L}{primo_dettaglio}:{L}{ultimo_dettaglio})"
            cella_num(r_fam, col, f, stile="fam")
        righe[fam] = r_fam

        if fam == "B":
            for col in range(3, 16):
                L = get_column_letter(col)
                cella_num(r, col, f"={L}{righe['A']}+{L}{righe['B']}", stile="tot")
            ws.cell(row=r, column=2, value="MARGINE DI CONTRIBUZIONE").font = F_TOT
            ws.cell(row=r, column=2).fill = FILL_TOT
            ws.cell(row=r, column=2).border = BORDO_TOP
            ws.cell(row=r, column=1).fill = FILL_TOT
            ws.cell(row=r, column=1).border = BORDO_TOP
            righe["MDC"] = r
            r += 2
        elif fam == "F":
            parti = "+".join(f"{{L}}{righe[k]}" for k in "ABCDEF")
            for col in range(3, 16):
                L = get_column_letter(col)
                cella_num(r, col, "=" + parti.replace("{L}", L), stile="tot")
            ws.cell(row=r, column=2, value="CASH FLOW OPERATIVO").font = F_TOT
            ws.cell(row=r, column=2).fill = FILL_TOT
            ws.cell(row=r, column=2).border = BORDO_TOP
            ws.cell(row=r, column=1).fill = FILL_TOT
            ws.cell(row=r, column=1).border = BORDO_TOP
            righe["CFO"] = r
            r += 2
        else:
            r += 1

    # variazione di cassa e saldi
    for col in range(3, 16):
        L = get_column_letter(col)
        f = f"={L}{righe['CFO']}+{L}{righe['G']}+{L}{righe['H']}+{L}{righe['I']}"
        cella_num(r, col, f, stile="tot")
    ws.cell(row=r, column=2, value="VARIAZIONE DI CASSA").font = F_TOT
    ws.cell(row=r, column=2).fill = FILL_TOT
    ws.cell(row=r, column=2).border = BORDO_TOP
    ws.cell(row=r, column=1).fill = FILL_TOT
    ws.cell(row=r, column=1).border = BORDO_TOP
    righe["VAR"] = r
    r += 1

    r_iniz = r
    ws.cell(row=r, column=2, value="Cassa e banca a inizio mese").font = F_SOTTO
    c = cella_num(r, 3, stile="sotto")
    c.fill = FILL_INPUT
    for col in range(4, 15):
        L_prec = get_column_letter(col - 1)
        cella_num(r, col, f"={L_prec}{r + 1}", stile="sotto")
    cella_num(r, 15, f"=C{r}", stile="sotto")
    r += 1

    for col in range(3, 16):
        L = get_column_letter(col)
        cella_num(r, col, f"={L}{r_iniz}+{L}{righe['VAR']}", stile="tot")
    ws.cell(row=r, column=2, value="CASSA E BANCA A FINE MESE").font = F_TOT
    ws.cell(row=r, column=2).fill = FILL_TOT
    ws.cell(row=r, column=2).border = BORDO_TOP
    ws.cell(row=r, column=1).fill = FILL_TOT
    ws.cell(row=r, column=1).border = BORDO_TOP
    cella_num(r, 15, f"=N{r}", stile="tot")
    righe["CASSA_FIN"] = r
    righe["CASSA_INIZ"] = r_iniz
    r += 3

    # ------------------------------------------------------------ memo
    ws.cell(row=r, column=2, value="RIGHE MEMO — fuori da ogni totale").font = F_MEMO_B
    r += 1
    memo_def = [
        ("Totale manodopera", "28 · 29 · 26.03 · 26.05 · 26.06 · 40.3.03",
         lambda L: f"={L}{righe['C']}+{L}{righe['D2']}+{L}{righe['G5']}"),
        ("Margine eventi", "ricavi eventi meno costi diretti eventi",
         lambda L: f"={L}{righe['A2']}+{L}{righe['D']}"),
        ("Tesoreria interna", "40.4 versamenti e giroconti: devono elidersi", None),
    ]
    for nome, spiega, f in memo_def:
        ws.cell(row=r, column=1, value="memo").font = F_MEMO
        ws.cell(row=r, column=2, value=f"   {nome} — {spiega}").font = F_MEMO
        for col in range(3, 16):
            L = get_column_letter(col)
            c = ws.cell(row=r, column=col)
            if f:
                c.value = f(L)
            else:
                c.fill = FILL_INPUT
            c.number_format = NUM
            c.font = F_MEMO
        if not f:
            ws.cell(row=r, column=15).value = f"=SUM(C{r}:N{r})"
        righe["MEMO_" + nome.split()[0]] = r
        r += 1
    r += 2

    # -------------------------------------------------------- controlli
    ws.cell(row=r, column=2, value="CONTROLLI DI QUADRATURA").font = F_MEMO_B
    r += 1
    controlli = [
        ("Variazione del prospetto meno variazione reale dei saldi",
         "deve essere zero: se non lo e', ci sono voci non mappate"),
        ("Versamenti contanti: gamba banca piu' gamba cassa",
         "deve essere zero: le due gambe si elidono"),
        ("Movimenti senza voce di conto (numero)",
         "deve essere zero: righe non categorizzate spariscono dal prospetto"),
        ("Movimenti su conti fuori piano o inattivi (numero)",
         "deve essere zero: etichette duplicate o conti legacy"),
    ]
    for i, (nome, spiega) in enumerate(controlli, 1):
        ws.cell(row=r, column=1, value=f"C{i}").font = F_MEMO
        ws.cell(row=r, column=2, value=f"   {nome}").font = F_MEMO
        for col in range(3, 16):
            c = ws.cell(row=r, column=col)
            c.number_format = NUM
            c.font = F_MEMO
            c.fill = FILL_INPUT
        ws.cell(row=r, column=15).value = f"=SUM(C{r}:N{r})"
        ws.cell(row=r, column=16, value=spiega).font = F_MEMO
        r += 1

    return righe


# ------------------------------------------------------------- SCOSTAMENTO
def foglio_scostamento(wb, righe_cf):
    ws = wb.create_sheet("SCOSTAMENTO")
    ws["A1"] = "SCOSTAMENTO — consuntivo contro budget"
    ws["A1"].font = F_TITOLO
    ws["A2"] = ("Differenza in euro e in percentuale. Le righe rispecchiano una per una "
                "quelle di CASH FLOW e BUDGET: stessa altezza, stessi codici.")
    ws["A2"].font = F_SOTTOTIT

    intest = ["Codice", "Voce"] + MESI + ["TOTALE", "", "Scost. %"]
    for i, h in enumerate(intest, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font, c.fill = F_HEADER, FILL_HEADER
        c.alignment = Alignment(horizontal="center" if i > 2 else "left", vertical="center")
    ws.row_dimensions[4].height = 22
    ws.freeze_panes = "C5"
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 46
    for i in range(3, 18):
        ws.column_dimensions[get_column_letter(i)].width = 13

    # rispecchia la geometria del foglio CASH FLOW riga per riga
    ultima = max(righe_cf.values()) + 1
    for r in range(5, ultima + 1):
        ws.cell(row=r, column=1).value = f"='CASH FLOW'!A{r}"
        ws.cell(row=r, column=2).value = f"='CASH FLOW'!B{r}"
        ws.cell(row=r, column=1).font = F_VOCE
        ws.cell(row=r, column=2).font = F_VOCE
        for col in range(3, 16):
            L = get_column_letter(col)
            c = ws.cell(row=r, column=col)
            c.value = f"=IF(AND('CASH FLOW'!{L}{r}=0,BUDGET!{L}{r}=0),\"\",'CASH FLOW'!{L}{r}-BUDGET!{L}{r})"
            c.number_format = NUM
            c.font = F_VOCE
        c = ws.cell(row=r, column=17)
        c.value = f"=IF(BUDGET!O{r}=0,\"\",('CASH FLOW'!O{r}-BUDGET!O{r})/ABS(BUDGET!O{r}))"
        c.number_format = PCT
        c.font = F_VOCE
    return ws


# ---------------------------------------------------------------- PER CENTRO
def foglio_per_centro(wb):
    ws = wb.create_sheet("PER CENTRO")
    ws["A1"] = "LETTURA PER CENTRO DI COSTO — totali dell'anno"
    ws["A1"].font = F_TITOLO
    ws["A2"] = ("Il centro segue l'attivita', non il conto corrente. I costi comuni restano su STR: "
                "il ribaltamento in fondo e' un calcolo di analisi, non scrive movimenti.")
    ws["A2"].font = F_SOTTOTIT

    intest = ["Codice", "Voce", "STR", "WEISS", "VV", "CAS", "TOTALE", "", "Quadratura"]
    for i, h in enumerate(intest, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font, c.fill = F_HEADER, FILL_HEADER
        c.alignment = Alignment(horizontal="center" if i > 2 else "left", vertical="center")
    ws.row_dimensions[4].height = 22
    ws.freeze_panes = "C5"
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 46
    for col, w in zip("CDEFGHI", [14, 14, 14, 14, 15, 3, 26]):
        ws.column_dimensions[col].width = w

    r = 5
    righe = {}
    for fam, fam_nome, sottogruppi in STRUTTURA:
        r_fam = r
        for col, val in ((1, fam), (2, fam_nome)):
            c = ws.cell(row=r, column=col, value=val)
            c.font, c.fill = F_FAM, FILL_FAM
        r += 1
        da = r
        for sot, sot_nome, _ in sottogruppi:
            ws.cell(row=r, column=1, value=sot).font = F_SOTTO
            ws.cell(row=r, column=2, value="   " + sot_nome).font = F_SOTTO
            for col in range(3, 7):
                c = ws.cell(row=r, column=col)
                c.number_format, c.font, c.fill = NUM, F_SOTTO, FILL_INPUT
            c = ws.cell(row=r, column=7, value=f"=SUM(C{r}:F{r})")
            c.number_format, c.font = NUM, F_SOTTO
            righe[sot] = r
            r += 1
        a = r - 1
        for col in range(3, 8):
            L = get_column_letter(col)
            c = ws.cell(row=r_fam, column=col, value=f"=SUM({L}{da}:{L}{a})")
            c.number_format, c.font, c.fill = NUM, F_FAM, FILL_FAM
        c = ws.cell(row=r_fam, column=9,
                    value=f"=IF(G{r_fam}='CASH FLOW'!O{{}},\"ok\",\"non torna\")")
        c.font = F_MEMO
        righe[fam] = r_fam
        r += 1

    r += 2
    ws.cell(row=r, column=2, value="RIBALTAMENTO DEI COSTI COMUNI (STR)").font = F_MEMO_B
    r += 1
    ws.cell(row=r, column=2,
            value="   Chiavi percentuali modificabili. Ogni riga deve sommare al 100%.").font = F_MEMO
    r += 1
    for i, h in enumerate(["Chiave", "Base di riparto", "WEISS %", "VV %", "CAS %", "Totale"], 1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = F_SOTTO
    r += 1
    chiavi = [
        ("Generale", "quota decisa a tavolino, default", 0.60, 0.30, 0.10),
        ("Per ricavi", "peso dei corrispettivi di ciascun centro", None, None, None),
        ("Per personale", "peso del costo del personale operativo", None, None, None),
        ("Per superficie", "metri quadri occupati, utile sugli immobili", None, None, None),
    ]
    for nome, base, w, v, cas in chiavi:
        ws.cell(row=r, column=1, value=nome).font = F_VOCE
        ws.cell(row=r, column=2, value=base).font = F_VOCE
        for col, val in ((3, w), (4, v), (5, cas)):
            c = ws.cell(row=r, column=col, value=val)
            c.number_format, c.font, c.fill = PCT, F_VOCE, FILL_INPUT
        c = ws.cell(row=r, column=6, value=f"=SUM(C{r}:E{r})")
        c.number_format = PCT
        c.font = F_SOTTO
        r += 1
    return ws, righe


# -------------------------------------------------------------------- SINTESI
def foglio_sintesi(wb, cf):
    ws = wb.create_sheet("SINTESI")
    ws["A1"] = "SINTESI — indicatori dell'anno"
    ws["A1"].font = F_TITOLO
    ws["A2"] = "Tutti gli indicatori leggono la colonna TOTALE del foglio CASH FLOW."
    ws["A2"].font = F_SOTTOTIT
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 62

    T = "'CASH FLOW'!O"
    kpi = [
        ("VOLUMI", None, None, None),
        ("Incassi operativi", f"={T}{cf['A']}", NUM, "Corrispettivi, eventi e altri proventi, al netto IVA"),
        ("Giorni di apertura", None, '0', "Da compilare: serve alla media giornaliera"),
        ("Incasso medio giornaliero", "=IF(C6=0,\"\",C5/C6)", NUM, "Incassi operativi diviso giorni di apertura"),
        (None, None, None, None),
        ("MARGINI", None, None, None),
        ("Margine di contribuzione", f"={T}{cf['MDC']}", NUM, "Incassi meno costo del venduto"),
        ("Margine di contribuzione %", f"=IF({T}{cf['A']}=0,\"\",{T}{cf['MDC']}/{T}{cf['A']})", PCT, None),
        ("Incidenza costo del venduto", f"=IF({T}{cf['A']}=0,\"\",-{T}{cf['B']}/{T}{cf['A']})", PCT,
         "Beverage, caffetteria, food e consumabili sugli incassi"),
        ("Incidenza manodopera completa", f"=IF({T}{cf['A']}=0,\"\",-{T}{cf['MEMO_Totale']}/{T}{cf['A']})", PCT,
         "Dalla riga memo: comprende personale eventi, ritenute e contributi"),
        ("Incidenza commissioni su incassi", f"=IF({T}{cf['A1']}=0,\"\",-{T}{cf['F3']}/{T}{cf['A1']})", PCT,
         "Pagobancomat, carte, SumUp, Axerve sui corrispettivi"),
        (None, None, None, None),
        ("CASH FLOW", None, None, None),
        ("Cash flow operativo", f"={T}{cf['CFO']}", NUM, "Prima di fisco, investimenti e finanziamenti"),
        ("Cash flow operativo %", f"=IF({T}{cf['A']}=0,\"\",{T}{cf['CFO']}/{T}{cf['A']})", PCT, None),
        ("Variazione di cassa", f"={T}{cf['VAR']}", NUM, None),
        ("Cassa e banca a inizio anno", f"='CASH FLOW'!C{cf['CASSA_INIZ']}", NUM, None),
        ("Cassa e banca a fine anno", f"='CASH FLOW'!N{cf['CASSA_FIN']}", NUM, None),
        (None, None, None, None),
        ("EVENTI", None, None, None),
        ("Ricavi eventi", f"={T}{cf['A2']}", NUM, None),
        ("Margine eventi", f"={T}{cf['MEMO_Margine']}", NUM, "Ricavi eventi meno costi diretti eventi"),
        ("Peso eventi sugli incassi", f"=IF({T}{cf['A']}=0,\"\",{T}{cf['A2']}/{T}{cf['A']})", PCT, None),
        (None, None, None, None),
        ("TENUTA", None, None, None),
        ("Costi di struttura mensili medi", f"=-{T}{cf['E']}/12", NUM, "Media dei dodici mesi"),
        ("Mesi di struttura coperti dalla cassa", None, '0.0', "Cassa finale diviso costi di struttura mensili"),
    ]

    r = 4
    for nome, formula, fmt, nota in kpi:
        if nome is None:
            r += 1
            continue
        if formula is None and fmt is None:
            c = ws.cell(row=r, column=2, value=nome)
            c.font = F_SOTTO
            c.fill = FILL_FAM
            ws.cell(row=r, column=3).fill = FILL_FAM
            r += 1
            continue
        ws.cell(row=r, column=2, value=nome).font = F_VOCE
        c = ws.cell(row=r, column=3)
        if formula:
            c.value = formula
        else:
            c.fill = FILL_INPUT
        c.number_format = fmt
        c.font = F_SOTTO
        if nota:
            ws.cell(row=r, column=4, value=nota).font = F_MEMO
        r += 1

    # dipende dalla riga dei costi di struttura, scritta poco sopra
    ws.cell(row=r - 1, column=3).value = f"=IF(C{r - 2}=0,\"\",C21/C{r - 2})"
    ws.cell(row=r - 1, column=3).number_format = "0.0"
    return ws


# ----------------------------------------------------------------- JSON
TOTALI = [
    {"codice": "MDC", "nome": "Margine di contribuzione", "somma": ["A", "B"]},
    {"codice": "CFO", "nome": "Cash flow operativo", "somma": ["A", "B", "C", "D", "E", "F"]},
    {"codice": "VAR", "nome": "Variazione di cassa", "somma": ["CFO", "G", "H", "I"]},
    {"codice": "CASSA_FIN", "nome": "Cassa e banca a fine periodo",
     "somma": ["CASSA_INIZ", "VAR"]},
]

RIGHE_MEMO = [
    {"codice": "M1", "nome": "Totale manodopera",
     "somma": ["C", "D2", "G5"],
     "scopo": "Percentuale manodopera sugli incassi, comprensiva di personale eventi, "
              "ritenute e contributi"},
    {"codice": "M2", "nome": "Margine eventi", "somma": ["A2", "D"],
     "scopo": "Verifica se gli eventi guadagnano"},
    {"codice": "M3", "nome": "Tesoreria interna", "voci": ["40.4.01", "40.4.02"],
     "scopo": "Versamenti e giroconti: si elidono nel consolidato, vanno sorvegliati"},
]

CONTROLLI = [
    {"codice": "C1", "nome": "Quadratura col saldo reale",
     "regola": "somma del prospetto = variazione dei saldi di cassa e banca nel periodo",
     "atteso": 0, "intercetta": "voci non mappate, movimenti persi"},
    {"codice": "C2", "nome": "Versamenti contanti a due gambe",
     "regola": "40.4.01 in banca + 40.4.01 in cassa = 0",
     "atteso": 0, "intercetta": "versamenti registrati su una gamba sola"},
    {"codice": "C3", "nome": "Movimenti senza voce di conto",
     "regola": "conteggio dei movimenti con accountId nullo", "atteso": 0,
     "intercetta": "righe non categorizzate che spariscono dal prospetto"},
    {"codice": "C4", "nome": "Movimenti su conti fuori piano",
     "regola": "conteggio dei movimenti su conti inattivi o non presenti nel piano",
     "atteso": 0, "intercetta": "etichette duplicate, conti legacy"},
]

CENTRI = [
    {"codice": "STR", "nome": "Struttura / Amministrazione", "tipo": "Comune"},
    {"codice": "WEISS", "nome": "Weiss Cafe'", "tipo": "Operativo"},
    {"codice": "VV", "nome": "Villa Varda Bistrot", "tipo": "Operativo"},
    {"codice": "CAS", "nome": "Casetta", "tipo": "Operativo stagionale"},
]


def scrivi_json(voci, ordine, path):
    import json

    famiglie = []
    for i, (fam, fam_nome, sottogruppi) in enumerate(STRUTTURA):
        sg = []
        for j, (sot, sot_nome, codici) in enumerate(sottogruppi):
            sg.append({
                "codice": sot,
                "nome": sot_nome,
                "ordine": j,
                "calcolato": not codici,
                "fonte": ("somma del campo vatAmount dei movimenti in "
                          + ("entrata" if sot == "G1" else "uscita")) if not codici else None,
                "voci": codici,
            })
        famiglie.append({
            "codice": fam, "nome": fam_nome, "ordine": i,
            "tipo": "RICAVO" if fam == "A" else (
                "PATRIMONIALE" if fam in ("G", "H", "I") else "COSTO"),
            "sottogruppi": sg,
        })

    payload = {
        "versione": "1.0",
        "data": "2026-08-11",
        "titolo": "Riclassificazione del piano dei conti v4 sul cash flow",
        "spec": "docs/superpowers/specs/2026-08-11-riclassificazione-cash-flow-design.md",
        "principio": "cassa",
        "convenzioneSegno": "entrate positive, uscite negative; i totali sono somme semplici",
        "regole": [
            "Un movimento si registra sempre sulla voce di dettaglio: famiglia e "
            "sottogruppo sono derivati, mai selezionabili a mano.",
            "Un conto appartiene a un solo sottogruppo.",
            "Le righe memo non sono categorie: attraversano famiglie diverse e non "
            "vanno modellate come BudgetCategory, altrimenti il totale conta due volte.",
            "Le righe G1 e G2 sono aggregazioni del campo vatAmount, non conti: "
            "nessun Account corrispondente.",
            "La natura la da' la voce, il luogo lo da' il centro di costo. Nessun "
            "sottogruppo contiene un riferimento a un locale.",
        ],
        "conteggi": {
            "vociTotali": len(ordine),
            "vociNelCashFlow": sum(1 for v in voci.values() if v["cassa"] == "SI"),
            "vociFuoriCassa": sum(1 for v in voci.values() if v["cassa"] == "NO"),
            "vociNelProspetto": sum(len(c) for _, _, sg in STRUTTURA for _, _, c in sg),
            "famiglie": len(STRUTTURA),
            "sottogruppi": sum(len(sg) for _, _, sg in STRUTTURA),
        },
        "centriDiCosto": CENTRI,
        "famiglie": famiglie,
        "totali": TOTALI,
        "memo": RIGHE_MEMO,
        "controlli": CONTROLLI,
        "voci": [
            {
                "codice": v["codice"],
                "nome": v["nome"],
                "sezione": v["sezione"],
                "mastro": v["mastro"],
                "gruppo": v["gruppo"] or None,
                "famiglia": v["famiglia"] or None,
                "sottogruppo": v["sottogruppo"] or None,
                "segno": v["segno"],
                "cassa": v["cassa"] == "SI",
                "regolaCentro": v["regola"],
                "nuova": v["codice"].startswith("40."),
                "note": v["note"] or None,
            }
            for c in ordine for v in [voci[c]]
        ],
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return payload


def main():
    voci, ordine = leggi_piano()
    voci = arricchisci(voci)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    foglio_mappatura(wb, voci, ordine)

    ws_cf = wb.create_sheet("CASH FLOW")
    righe_cf = scrivi_prospetto(
        ws_cf, voci, "CASH FLOW — consuntivo",
        "Solo movimenti che toccano cassa o banca. Entrate positive, uscite negative. "
        "Le voci di dettaglio sono nascoste: usa i pulsanti + a sinistra per aprirle.")

    ws_bd = wb.create_sheet("BUDGET")
    scrivi_prospetto(
        ws_bd, voci, "BUDGET — obiettivo",
        "Stessa struttura del consuntivo, riga per riga. Compila le voci di dettaglio: "
        "sottogruppi, famiglie e margini si calcolano da soli.")

    foglio_scostamento(wb, righe_cf)
    ws_pc, righe_pc = foglio_per_centro(wb)
    for fam, _, _ in STRUTTURA:
        ws_pc.cell(row=righe_pc[fam], column=9).value = (
            f"=IF(ROUND(G{righe_pc[fam]}-'CASH FLOW'!O{righe_cf[fam]},2)=0,\"ok\",\"non torna\")")
    foglio_sintesi(wb, righe_cf)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    wb.save(DST)
    JSON = "/Users/nicolascarpa/Desktop/accounting/docs/cash-flow-riclassificazione.json"
    scrivi_json(voci, ordine, JSON)
    print(f"scritto: {JSON}")
    print(f"scritto: {DST}")
    print(f"voci totali: {len(ordine)}")
    print(f"fuori cash flow: {sum(1 for v in voci.values() if v['cassa'] == 'NO')}")
    nel_prospetto = sum(len(c) for _, _, sg in STRUTTURA for _, _, c in sg)
    print(f"voci nel prospetto: {nel_prospetto}")
    print(f"sottogruppi: {sum(len(sg) for _, _, sg in STRUTTURA)}")
    orfane = [c for c, v in voci.items()
              if v["cassa"] == "SI" and not v["famiglia"]]
    print(f"voci orfane (cassa SI senza famiglia): {orfane}")


if __name__ == "__main__":
    main()
